"""
Script de Importacion de Catalogo
Importa productos desde Catalogo.xlsx a Supabase
"""

import openpyxl
import re
import requests
from datetime import datetime
import os
from dotenv import load_dotenv

load_dotenv()

# Configuracion
SUPABASE_URL = 'https://gyonguqndcsmudqmptfb.supabase.co'
SUPABASE_KEY = os.getenv('SUPABASE_KEY')

# Configuracion del archivo Excel
CONFIG = {
    'ARCHIVO': 'Catálogo.xlsx',
    'HOJA': 'Hoja 1',
    'HEADER_ROW': 7,
    'DATA_START_ROW': 8,
    'DOLAR_CELL': 'B5',
}

def supabase_request(method, endpoint, data=None):
    """Hacer peticion a Supabase usando requests"""
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal'
    }
    
    if method == 'GET':
        response = requests.get(url, headers=headers)
    elif method == 'POST':
        response = requests.post(url, headers=headers, json=data)
    elif method == 'PATCH':
        response = requests.patch(url, headers=headers, json=data)
    elif method == 'DELETE':
        response = requests.delete(url, headers=headers)
    
    if response.status_code not in [200, 201, 204]:
        raise Exception(f"Error en Supabase: {response.status_code} - {response.text}")
    
    return response

def parsear_especificaciones(espec_str):
    """Parsea 'GB - % bat.' para extraer almacenamiento y bateria"""
    if not espec_str:
        return None, None
    
    espec_str = str(espec_str).strip()
    match = re.search(r'(\d+\s*GB)\s*/\s*(\d+%)', espec_str, re.IGNORECASE)
    
    if match:
        return match.group(1).strip(), match.group(2).strip()
    
    match_gb = re.search(r'(\d+\s*GB)', espec_str, re.IGNORECASE)
    if match_gb:
        return match_gb.group(1).strip(), None
    
    return None, None

def calcular_precio(cell_value):
    """Calcula el precio"""
    if cell_value is None:
        return None
    
    if isinstance(cell_value, (int, float)):
        return round(float(cell_value), 2)
    
    try:
        return round(float(cell_value), 2)
    except:
        return None

def importar_catalogo():
    """Funcion principal de importacion"""
    print("=" * 80)
    print("IMPORTACION DE CATALOGO")
    print("=" * 80)
    
    # Cargar archivo Excel
    print(f"\n[*] Cargando archivo: {CONFIG['ARCHIVO']}")
    try:
        wb = openpyxl.load_workbook(CONFIG['ARCHIVO'], data_only=True)
        ws = wb.active
        print(f"[OK] Archivo cargado: {ws.max_row} filas, {ws.max_column} columnas")
    except Exception as e:
        print(f"[ERROR] Error al cargar archivo: {e}")
        return
    
    # Obtener cotizacion del dolar
    dolar_cell = ws[CONFIG['DOLAR_CELL']]
    cotizacion_dolar = dolar_cell.value
    
    if isinstance(cotizacion_dolar, str) and cotizacion_dolar.startswith('='):
        cotizacion_dolar = 1485.0
        print(f"\n[*] Cotizacion del dolar (por defecto): ${cotizacion_dolar}")
    else:
        try:
            cotizacion_dolar = float(cotizacion_dolar)
            print(f"\n[*] Cotizacion del dolar: ${cotizacion_dolar}")
        except:
            cotizacion_dolar = 1485.0
            print(f"\n[*] Cotizacion del dolar (por defecto): ${cotizacion_dolar}")
    
    # Procesar productos
    productos = []
    productos_procesados = 0
    productos_saltados = 0
    
    print(f"\n[*] Procesando productos desde fila {CONFIG['DATA_START_ROW']}...")
    
    for row_idx in range(CONFIG['DATA_START_ROW'], ws.max_row + 1):
        modelo = ws[f'A{row_idx}'].value
        colores = ws[f'B{row_idx}'].value
        especificaciones = ws[f'C{row_idx}'].value
        precio_usd = ws[f'D{row_idx}'].value
        
        if not modelo or not precio_usd:
            productos_saltados += 1
            continue
        
        almacenamiento, bateria = parsear_especificaciones(especificaciones)
        precio_usd_num = calcular_precio(precio_usd)
        
        if not precio_usd_num:
            productos_saltados += 1
            continue
        
        precio_ars = round(precio_usd_num * cotizacion_dolar, 2)
        cuotas_3 = round((precio_ars * 1.22) / 3, -3)
        cuotas_6 = round((precio_ars * 1.33) / 6, -3)
        cuotas_12 = round((precio_ars * 1.6) / 12, 2)
        
        producto = {
            'modelo': str(modelo).strip(),
            'colores': str(colores).strip() if colores else None,
            'almacenamiento': almacenamiento,
            'bateria': bateria,
            'precio_usd': precio_usd_num,
            'precio_ars': precio_ars,
            'cuotas_3': cuotas_3,
            'cuotas_6': cuotas_6,
            'cuotas_12': cuotas_12,
            'stock': 1,
            'activo': True
        }
        
        productos.append(producto)
        productos_procesados += 1
        
        if productos_procesados % 100 == 0:
            print(f"   Procesados: {productos_procesados} productos...")
    
    print(f"\n[OK] Productos procesados: {productos_procesados}")
    print(f"[!] Productos saltados: {productos_saltados}")
    
    # Insertar en Supabase
    if productos:
        print(f"\n[*] Insertando {len(productos)} productos en Supabase...")
        
        try:
            print("   Marcando productos existentes como inactivos...")
            supabase_request('PATCH', 'productos?activo=eq.true', {'activo': False})
            
            batch_size = 50
            productos_insertados = 0
            
            for i in range(0, len(productos), batch_size):
                batch = productos[i:i + batch_size]
                supabase_request('POST', 'productos', batch)
                productos_insertados += len(batch)
                print(f"   Insertados: {productos_insertados}/{len(productos)}")
            
            importacion = {
                'archivo_nombre': CONFIG['ARCHIVO'],
                'productos_importados': len(productos),
                'productos_actualizados': 0,
                'productos_nuevos': len(productos),
                'cotizacion_dolar': cotizacion_dolar,
                'estado': 'completado'
            }
            
            supabase_request('POST', 'importaciones_catalogo', importacion)
            
            print(f"\n[OK] Importacion completada exitosamente!")
            print(f"   - {len(productos)} productos importados")
            print(f"   - Cotizacion del dolar: ${cotizacion_dolar}")
            
        except Exception as e:
            print(f"\n[ERROR] Error al insertar en Supabase: {e}")
            
            importacion_error = {
                'archivo_nombre': CONFIG['ARCHIVO'],
                'productos_importados': 0,
                'cotizacion_dolar': cotizacion_dolar,
                'estado': 'error',
                'errores': {'error': str(e)}
            }
            
            try:
                supabase_request('POST', 'importaciones_catalogo', importacion_error)
            except:
                pass
    else:
        print("\n[!] No se encontraron productos para importar")
    
    print("\n" + "=" * 80)
    print("IMPORTACION FINALIZADA")
    print("=" * 80)

if __name__ == "__main__":
    importar_catalogo()
