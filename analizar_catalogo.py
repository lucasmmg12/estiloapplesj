import openpyxl
import json

# Cargar el archivo Excel
wb = openpyxl.load_workbook('Catálogo.xlsx')

output = []

output.append("=" * 80)
output.append("ANÁLISIS DEL ARCHIVO CATÁLOGO.XLSX")
output.append("=" * 80)

# Mostrar todas las hojas
output.append(f"\n📋 HOJAS DISPONIBLES: {wb.sheetnames}")

# Analizar la primera hoja (la más importante)
ws = wb.active
sheet_name = wb.active.title

output.append(f"\n{'=' * 80}")
output.append(f"HOJA ACTIVA: {sheet_name}")
output.append(f"{'=' * 80}")

# Dimensiones
output.append(f"\n📊 DIMENSIONES:")
output.append(f"   - Total de filas: {ws.max_row}")
output.append(f"   - Total de columnas: {ws.max_column}")

# Buscar la fila de encabezados (puede no ser la fila 1)
header_row = 1
headers = []

# Intentar encontrar encabezados en las primeras 10 filas
for row_idx in range(1, min(11, ws.max_row + 1)):
    row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
    non_none_count = sum(1 for v in row_values if v is not None and str(v).strip())
    
    if non_none_count >= ws.max_column * 0.5:  # Al menos 50% de columnas con datos
        header_row = row_idx
        headers = row_values
        break

output.append(f"\n📌 FILA DE ENCABEZADOS DETECTADA: Fila {header_row}")
output.append(f"\n📌 ENCABEZADOS:")
for idx, header in enumerate(headers, 1):
    col_letter = openpyxl.utils.get_column_letter(idx)
    output.append(f"   {idx}. Columna {col_letter}: '{header}'")

# Primeras 10 filas de datos DESPUÉS de los encabezados
data_start_row = header_row + 1
output.append(f"\n📝 PRIMERAS 10 FILAS DE DATOS (desde fila {data_start_row}):")

for row_idx in range(data_start_row, min(data_start_row + 10, ws.max_row + 1)):
    output.append(f"\n   === Fila {row_idx} ===")
    row_data = {}
    has_data = False
    
    for col_idx in range(1, min(ws.max_column + 1, 24)):  # Limitar a 23 columnas
        cell_value = ws.cell(row_idx, col_idx).value
        col_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else f"Col{col_idx}"
        
        if cell_value is not None:
            has_data = True
            output.append(f"      {col_name}: {cell_value}")
    
    if not has_data:
        output.append(f"      (fila vacía)")

# Guardar en archivo
with open('analisis_catalogo.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print("Análisis guardado en: analisis_catalogo.txt")
print(f"\nResumen rápido:")
print(f"- Hojas: {len(wb.sheetnames)}")
print(f"- Hoja activa: {sheet_name}")
print(f"- Filas totales: {ws.max_row}")
print(f"- Columnas totales: {ws.max_column}")
print(f"- Fila de encabezados: {header_row}")
print(f"- Encabezados encontrados: {len([h for h in headers if h is not None])}")
